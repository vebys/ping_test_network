from ping3 import ping
from openpyxl import load_workbook


def read_xls():
    # 获取文件对象
    wb2 = load_workbook('摄像头ip.xlsx')
    # 通过工作表名获取到工作表对象
    ws = wb2['Sheet1']
    i = 1
    fail_list = []
    success_list = []
    for row in ws.rows:
        ip = row[0].value
        print(ip)
        res = ping(ip)
        if res:
            success_list.append(ip)
            ws.cell(i, 2, 'ok')
        else:
            fail_list.append(ip)
            ws.cell(i, 2, 'fail')
        i += 1
        # if i > 100:
        #     break
    wb2.save('检测结果.xlsx')
    wb2.close()
    print('fail_list:', fail_list)
    print('success_list:', success_list)


if "__main__" == __name__:
    read_xls()
